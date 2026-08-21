# -*- coding: utf-8 -*-
"""
make_samples.py — 엑셀 업로드 양식 + 샘플 데이터 xlsx 생성기
실행: python3 make_samples.py  (openpyxl, node 필요)

생성 파일 (templates/):
  - marketing_template.xlsx : 빈 업로드 양식 (헤더 + 작성 안내)
  - marketing_sample.xlsx   : 샘플 데이터 채움 (js/sample-data.js 와 동일 내용)

샘플 값은 js/sample-data.js 를 node 로 읽어와 사용하므로 두 곳이 항상 일치한다.
"""
import json
import os
import subprocess

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(BASE, "templates")

NAVY = "0B2E59"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
NOTE_FONT = Font(color="68758C", size=9, italic=True)

# 시트 구성: (시트명, 헤더, 작성 안내, sample-data.js 키, 행 변환 함수)
SHEETS = [
    ("캠페인",
     ["ID", "캠페인명", "브랜드", "상태", "진행률(%)", "담당자", "시작일", "종료일", "다음 마일스톤"],
     "브랜드: Hyundai/Develon/공통 · 상태: 기획/진행/검토/완료/보류 · 날짜: YYYY-MM-DD",
     "campaigns",
     lambda c: [c["id"], c["name"], c["brand"], c["status"], c["progress"],
                c["owner"], c["start"], c["end"], c["milestone"]]),
    ("업무",
     ["ID", "업무명", "관련 캠페인", "담당자", "우선순위", "마감일", "상태"],
     "우선순위: 높음/중간/낮음 · 상태: 대기/진행/완료 · 마감일: YYYY-MM-DD",
     "tasks",
     lambda t: [t["id"], t["name"], t["campaign"], t["owner"], t["priority"], t["due"], t["status"]]),
    ("주간일정",
     ["ID", "날짜", "시간", "구분", "일정명"],
     "구분: 회의/마감/행사 · 날짜: YYYY-MM-DD · 시간: HH:MM (없으면 공란)",
     "events",
     lambda e: [e["id"], e["date"], e["time"], e["type"], e["title"]]),
    ("채널실적",
     ["월(YYYY-MM)", "채널", "값"],
     "채널: 뉴스레터/홈페이지/SNS/전시회·행사 등 · 값: 채널 대표 지표 수치",
     "channelStats",
     lambda r: [r["month"], r["channel"], r["value"]]),
    ("공지",
     ["ID", "고정(Y/N)", "중요도", "제목", "작성자", "날짜"],
     "고정: Y면 상단 고정 · 중요도: 높음/보통 · 날짜: YYYY-MM-DD",
     "notices",
     lambda n: [n["id"], "Y" if n["pinned"] else "N", n["importance"], n["title"], n["author"], n["date"]]),
]


def load_sample_data():
    """js/sample-data.js 를 node 로 읽어 dict 반환 (데이터 이중 관리 방지)"""
    out = subprocess.check_output(
        ["node", "-p", "JSON.stringify(require('./js/sample-data.js'))"],
        cwd=BASE)
    return json.loads(out.decode("utf-8"))


def build(filename, with_data, sample):
    wb = Workbook()
    wb.remove(wb.active)
    for name, headers, note, key, to_row in SHEETS:
        ws = wb.create_sheet(name)
        ws.append(headers)
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal="center", vertical="center")
        rows = [to_row(item) for item in sample[key]] if with_data else []
        for r in rows:
            ws.append(r)
        # 작성 안내 (데이터 오른쪽 끝 다음 열)
        note_col = len(headers) + 2
        nc = ws.cell(row=1, column=note_col, value="※ " + note)
        nc.font = NOTE_FONT
        # 열 너비
        for i, h in enumerate(headers, start=1):
            width = len(str(h)) * 2 + 4
            for r in rows:
                v = r[i - 1]
                width = max(width, min(len(str(v)) * 1.7 + 4, 46))
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.freeze_panes = "A2"
    path = os.path.join(OUT_DIR, filename)
    wb.save(path)
    print("생성:", os.path.relpath(path, BASE))


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    sample = load_sample_data()
    build("marketing_template.xlsx", with_data=False, sample=sample)
    build("marketing_sample.xlsx", with_data=True, sample=sample)


if __name__ == "__main__":
    main()
